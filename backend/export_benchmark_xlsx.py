#!/usr/bin/env python3
"""
Export the benchmark JSON to a formatted .xlsx workbook.

Formatting:
  - Frozen header row + autofilter
  - Color-coded column groups (meta / CysticCare RAG / graders / GPT-5.4 / Sonnet 4.6)
  - Bold white header, wrapped text on long columns, top alignment, sensible widths
  - Numeric grader scores written as numbers; source lists rendered readable
  - A "Legend" sheet documenting the columns

Run:  /path/to/venv/bin/python backend/export_benchmark_xlsx.py \
        --json "benchmark result - benchmark result.json" \
        --xlsx "benchmark result - benchmark result.xlsx"
"""
import argparse
import json

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.dimensions import ColumnDimension

# desired column order
ORDER = [
    "question", "type", "category", "difficulty", "expected_behavior", "gold_answer",
    "cysticcare_response", "cysticcare_sources", "source_papers", "supporting_passages",
    "prompt_1_input", "prompt_1_Text similarity grader_score", "prompt_1_Text similarity grader_pass",
    "prompt_2_input", "prompt_2_Auto grader_score", "prompt_2_Auto grader_pass",
    "gpt5.4_response", "gpt5.4_sources", "sonnet4.6_response", "sonnet4.6_sources",
]

# group -> (header fill hex, columns)
GROUPS = [
    ("Question / metadata", "4472C4",
     ["question", "type", "category", "difficulty", "expected_behavior", "gold_answer"]),
    ("CysticCare RAG", "548235",
     ["cysticcare_response", "cysticcare_sources", "source_papers", "supporting_passages"]),
    ("Graders", "7F7F7F",
     ["prompt_1_input", "prompt_1_Text similarity grader_score", "prompt_1_Text similarity grader_pass",
      "prompt_2_input", "prompt_2_Auto grader_score", "prompt_2_Auto grader_pass"]),
    ("GPT-5.4 (web search)", "C55A11", ["gpt5.4_response", "gpt5.4_sources"]),
    ("Sonnet 4.6 (web search)", "7030A0", ["sonnet4.6_response", "sonnet4.6_sources"]),
]
COL_FILL = {c: hexcol for _, hexcol, cols in GROUPS for c in cols}

WIDE = {"question", "gold_answer", "cysticcare_response", "cysticcare_sources", "source_papers",
        "supporting_passages", "prompt_1_input", "prompt_2_input",
        "gpt5.4_response", "gpt5.4_sources", "sonnet4.6_response", "sonnet4.6_sources"}
WIDTHS = {"type": 18, "category": 22, "difficulty": 12, "expected_behavior": 26,
          "prompt_1_Text similarity grader_score": 13, "prompt_1_Text similarity grader_pass": 12,
          "prompt_2_Auto grader_score": 13, "prompt_2_Auto grader_pass": 12}
NUMERIC = {"prompt_1_Text similarity grader_score", "prompt_2_Auto grader_score"}
SOURCE_LISTS = {"gpt5.4_sources", "sonnet4.6_sources"}


def fmt_sources(v):
    if isinstance(v, list):
        return "\n".join(
            f"{i}. {(s.get('title') or '').strip()}\n   {s.get('url','')}"
            for i, s in enumerate(v, 1)
        )
    return "" if v is None else str(v)


def cell_value(col, raw):
    if col in SOURCE_LISTS:
        return fmt_sources(raw)
    if col in NUMERIC:
        try:
            return float(raw)
        except (TypeError, ValueError):
            return None
    return "" if raw is None else str(raw)


def build(json_path, xlsx_path):
    recs = json.load(open(json_path))
    wb = Workbook()
    ws = wb.active
    ws.title = "Benchmark"

    thin = Side(style="thin", color="D9D9D9")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    header_font = Font(bold=True, color="FFFFFF", size=11)
    top_wrap = Alignment(vertical="top", wrap_text=True)
    top_left = Alignment(vertical="top", horizontal="left")
    center = Alignment(vertical="center", horizontal="center", wrap_text=True)

    # header
    for j, col in enumerate(ORDER, 1):
        c = ws.cell(row=1, column=j, value=col)
        c.font = header_font
        c.fill = PatternFill("solid", fgColor=COL_FILL.get(col, "404040"))
        c.alignment = center
        c.border = border
        letter = get_column_letter(j)
        ws.column_dimensions[letter].width = 75 if col in WIDE else WIDTHS.get(col, 16)

    # data
    for i, rec in enumerate(recs, start=2):
        for j, col in enumerate(ORDER, 1):
            val = cell_value(col, rec.get(col))
            c = ws.cell(row=i, column=j, value=val)
            c.border = border
            if col in NUMERIC:
                c.alignment = top_left
                c.number_format = "0.000"
            elif col in WIDE or col in SOURCE_LISTS:
                c.alignment = top_wrap
            else:
                c.alignment = top_left

    last_col = get_column_letter(len(ORDER))
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{last_col}{len(recs) + 1}"
    ws.row_dimensions[1].height = 30
    ws.sheet_view.zoomScale = 100

    # ---- Legend sheet ----
    lg = wb.create_sheet("Legend")
    lg.column_dimensions["A"].width = 34
    lg.column_dimensions["B"].width = 90
    lg.append(["Column", "Meaning"])
    for c in lg[1]:
        c.font = Font(bold=True, color="FFFFFF"); c.fill = PatternFill("solid", fgColor="404040")
    notes = [
        ("— Question / metadata —", ""),
        ("question", "Benchmark question"),
        ("type / category / difficulty", "Question strata (this export focuses on the Standard factual type)"),
        ("expected_behavior", "Intended behavior (answer / abstain / correct_the_premise / ...)"),
        ("gold_answer", "Reference answer with [P#] citations"),
        ("— CysticCare RAG —", ""),
        ("cysticcare_response", "Your RAG system's answer"),
        ("cysticcare_sources", "Papers the RAG retrieved (with relevance scores)"),
        ("source_papers / supporting_passages", "Retrieved evidence used by the RAG"),
        ("— Graders —", ""),
        ("prompt_1_* (Text similarity)", "Semantic-similarity grader: score (0-1) + pass"),
        ("prompt_2_* (Auto grader)", "LLM rubric grader: score (0-1) + pass"),
        ("— Baselines (this task) —", ""),
        ("gpt5.4_response / _sources", "GPT-5.4 (OpenAI Responses API + web_search) answer and cited URLs"),
        ("sonnet4.6_response / _sources", "Claude Sonnet 4.6 (web_search_20250305, max_uses=2) answer and cited URLs"),
        ("", ""),
        ("Note", "Baseline fields are populated for Standard factual rows only. "
                 "In Excel, Select All → Format → AutoFit Row Height to expand wrapped cells."),
    ]
    for a, b in notes:
        r = lg.max_row + 1
        lg.cell(row=r, column=1, value=a).font = Font(bold=a.startswith("—") or a == "Note")
        lg.cell(row=r, column=2, value=b).alignment = Alignment(wrap_text=True, vertical="top")
    lg.freeze_panes = "A2"

    wb.save(xlsx_path)
    return len(recs), len(ORDER)


if __name__ == "__main__":
    ap = argparse.ArgumentParser()
    ap.add_argument("--json", required=True)
    ap.add_argument("--xlsx", required=True)
    a = ap.parse_args()
    n, cols = build(a.json, a.xlsx)
    print(f"Wrote {a.xlsx}: {n} data rows x {cols} columns (+ Legend sheet)")
